"""
Komplai Demo Pipeline - Excel Manager
Handles reading from and appending to Excel files
"""

import pandas as pd
from typing import List, Dict
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelManager:
    """Manage Excel file operations for the demo pipeline"""

    def __init__(self, excel_path: str):
        """
        Initialize Excel Manager

        Args:
            excel_path: Path to the Excel master file
        """
        self.excel_path = excel_path

        # Verify file exists
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Read a sheet from the Excel file

        Args:
            sheet_name: Name of the sheet to read

        Returns:
            DataFrame with the sheet data
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name=sheet_name)
            print(f"✓ Read {len(df)} rows from {sheet_name}")
            return df
        except Exception as e:
            print(f"✗ Error reading sheet {sheet_name}: {e}")
            raise

    def append_to_sheet(self, sheet_name: str, data: List[Dict]) -> bool:
        """
        Append rows to an existing sheet in the Excel file

        Args:
            sheet_name: Name of the sheet to append to
            data: List of dictionaries with row data

        Returns:
            True if successful, False otherwise
        """
        if not data:
            print(f"⚠ No data to append to {sheet_name}")
            return True

        try:
            # Convert list of dicts to DataFrame
            new_df = pd.DataFrame(data)

            # Load existing workbook
            book = load_workbook(self.excel_path)

            if sheet_name not in book.sheetnames:
                raise ValueError(f"Sheet {sheet_name} not found in workbook")

            # Get the sheet
            sheet = book[sheet_name]

            # Find the next empty row
            next_row = sheet.max_row + 1

            # Append rows (without header)
            for r_idx, row in enumerate(dataframe_to_rows(new_df, index=False, header=False), next_row):
                for c_idx, value in enumerate(row, 1):
                    # Handle datetime objects
                    if isinstance(value, pd.Timestamp):
                        value = value.to_pydatetime()
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Save the workbook
            book.save(self.excel_path)
            book.close()

            print(f"✓ Appended {len(data)} rows to {sheet_name}")
            return True

        except Exception as e:
            print(f"✗ Error appending to sheet {sheet_name}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def get_last_row(self, sheet_name: str) -> pd.Series:
        """
        Get the last row from a sheet

        Args:
            sheet_name: Name of the sheet

        Returns:
            Last row as a pandas Series
        """
        try:
            df = self.read_sheet(sheet_name)
            if len(df) > 0:
                return df.iloc[-1]
            else:
                return pd.Series()
        except Exception as e:
            print(f"✗ Error getting last row from {sheet_name}: {e}")
            return pd.Series()

    def get_all_sheets(self) -> List[str]:
        """
        Get list of all sheet names in the Excel file

        Returns:
            List of sheet names
        """
        try:
            xl = pd.ExcelFile(self.excel_path)
            return xl.sheet_names
        except Exception as e:
            print(f"✗ Error getting sheet names: {e}")
            return []

    def backup_file(self, backup_suffix: str = None) -> str:
        """
        Create a backup of the Excel file

        Args:
            backup_suffix: Optional suffix for backup file (default: timestamp)

        Returns:
            Path to the backup file
        """
        try:
            if backup_suffix is None:
                backup_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")

            base_path = os.path.splitext(self.excel_path)[0]
            backup_path = f"{base_path}_backup_{backup_suffix}.xlsx"

            # Copy file
            import shutil
            shutil.copy2(self.excel_path, backup_path)

            print(f"✓ Created backup: {backup_path}")
            return backup_path

        except Exception as e:
            print(f"✗ Error creating backup: {e}")
            return None


if __name__ == "__main__":
    # Test Excel Manager
    print("\n" + "="*60)
    print("TESTING EXCEL MANAGER")
    print("="*60)

    # Initialize
    excel_path = "Komplai_Demo_Master (Claude).xlsx"
    manager = ExcelManager(excel_path)

    # Test reading
    print("\n[Reading Sheets]")
    entities = manager.read_sheet('Entities')
    print(f"Entities: {len(entities)} rows, {len(entities.columns)} columns")

    invoices = manager.read_sheet('Invoices_Master')
    print(f"Invoices: {len(invoices)} rows, {len(invoices.columns)} columns")

    # Test getting last row
    print("\n[Last Row]")
    last_invoice = manager.get_last_row('Invoices_Master')
    if not last_invoice.empty:
        print(f"Last Invoice: {last_invoice.get('Invoice_ID', 'N/A')}")

    # Test getting all sheets
    print("\n[All Sheets]")
    sheets = manager.get_all_sheets()
    print(f"Sheets: {sheets}")

    print("\n" + "="*60)
    print("✓ EXCEL MANAGER TEST COMPLETE")
    print("="*60)
