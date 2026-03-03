from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import date, timedelta

wb = Workbook()

# ============================================================
# COLOR PALETTE
# ============================================================
NAVY = "1B2A4A"
ORANGE = "E85D26"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY = "D9D9D9"
DARK_GRAY = "333333"
GREEN_BG = "E2EFDA"
YELLOW_BG = "FFF2CC"
RED_BG = "FCE4EC"
BLUE_BG = "D6E4F0"
ORANGE_BG = "FDE9D9"
PHASE_0_COLOR = "E8EAF6"  # Indigo light
PHASE_1_COLOR = "E3F2FD"  # Blue light
PHASE_2_COLOR = "E8F5E9"  # Green light
PHASE_3_COLOR = "FFF3E0"  # Orange light

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

header_font = Font(name='Arial', bold=True, size=11, color=WHITE)
header_fill = PatternFill('solid', fgColor=NAVY)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

body_font = Font(name='Arial', size=10, color=DARK_GRAY)
body_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

phase_font = Font(name='Arial', bold=True, size=11, color=NAVY)
section_font = Font(name='Arial', bold=True, size=10, color=ORANGE)

# ============================================================
# SHEET 1: MASTER ACTION PLAN
# ============================================================
ws = wb.active
ws.title = "Master Action Plan"
ws.sheet_properties.tabColor = NAVY

# Column config: ID, Phase, Category, Task, Sub-tasks/Details, Owner, Priority, Dependencies, Start, End, Duration (days), Status, Success Metric, Notes
cols = {
    'A': ('ID', 5),
    'B': ('Phase', 14),
    'C': ('Category', 14),
    'D': ('Task', 40),
    'E': ('Sub-tasks / Details', 55),
    'F': ('Owner', 12),
    'G': ('Priority', 10),
    'H': ('Dependency', 12),
    'I': ('Start Date', 13),
    'J': ('End Date', 13),
    'K': ('Days', 7),
    'L': ('Status', 12),
    'M': ('Success Metric', 35),
    'N': ('Notes / Risk', 35),
}

for col_letter, (name, width) in cols.items():
    ws.column_dimensions[col_letter].width = width

# Title rows
ws.merge_cells('A1:N1')
ws['A1'] = "KOMPLAI | PROJECT HAIL MARY — MASTER ACTION PLAN"
ws['A1'].font = Font(name='Arial', bold=True, size=16, color=NAVY)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:N2')
ws['A2'] = "Larry Self-Serve Launch | Feb 17 – Sep 30, 2026 | Owners: Rishi (GTM/Strategy), Anirudh (Product/Engineering)"
ws['A2'].font = Font(name='Arial', size=10, color='666666')
ws['A2'].alignment = Alignment(horizontal='center')

# Headers in row 4
for col_letter, (name, _) in cols.items():
    cell = ws[f'{col_letter}4']
    cell.value = name
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.row_dimensions[4].height = 30
ws.freeze_panes = 'A5'
ws.auto_filter.ref = 'A4:N4'

# ============================================================
# TASK DATA
# ============================================================
# Each task: [phase, category, task, details, owner, priority, dependency, start, end, status, success_metric, notes]
# Dates: Feb 17 is Monday

tasks = [
    # ============================================================
    # PHASE 0: PRE-LAUNCH (Feb 17 - Mar 31)
    # ============================================================
    # --- WEEK 1: Feb 17-21 ---
    ["PHASE 0", "", "WEEK 1: Feb 17-21 — Foundation Sprint", "", "", "", "", "", "", "", "", ""],

    ["Phase 0: Pre-Launch", "Product", "Self-serve onboarding flow: end-to-end audit",
     "Map every screen from signup → ERP connect → first Larry insight. Document every click, every load time, every error state. Record screen captures of the full flow for all 4 ERPs (Zoho, QBO, ERPNext, Xero).",
     "Anirudh", "P0 - Critical", "None",
     "2026-02-17", "2026-02-19", "Not Started",
     "Full flow documented with screenshots for all 4 ERPs", "This is THE most important task. Nothing else matters if onboarding is broken."],

    ["Phase 0: Pre-Launch", "Product", "Build pre-populated first queries for Larry",
     "After ERP sync, Larry should suggest 3-5 data-specific queries (not generic). If user has overdue receivables → suggest aging analysis. If burn spiked → suggest burn analysis. Requires conditional logic based on synced data.",
     "Anirudh", "P0 - Critical", "Onboarding audit complete",
     "2026-02-19", "2026-02-28", "Not Started",
     "Larry suggests personalized queries within 60 seconds of ERP sync", "The difference between 'ask Larry anything' and 'Larry already found something' is the difference between activation and abandonment."],

    ["Phase 0: Pre-Launch", "Product", "Build 'wow moment' automated email",
     "Triggered 15-30 min after ERP sync completes. Email must contain 2-3 REAL insights from the user's actual data. E.g., 'Larry found $47K in receivables over 90 days across 10 invoices.' Requires email template + dynamic data pipeline.",
     "Anirudh", "P0 - Critical", "Pre-populated queries built",
     "2026-02-24", "2026-03-05", "Not Started",
     "Email sends within 30 min of ERP sync with real data points", "This email is the single most important piece of marketing you will produce."],

    ["Phase 0: Pre-Launch", "Product", "Build guided product walkthrough (3-step tour)",
     "Interactive overlay (not video): Step 1 → first query, Step 2 → first chart, Step 3 → attempt download (gates on Scaling tier). Must complete in under 2 minutes. Use product tour library (e.g., Shepherd.js or similar).",
     "Anirudh", "P1 - High", "Onboarding audit complete",
     "2026-02-20", "2026-03-03", "Not Started",
     "Tour completion rate >60% of new users", "Keep it short. 3 steps max. Users who complete onboarding tours convert 2-3x better."],

    ["Phase 0: Pre-Launch", "Product", "Build error handling for ERP sync failures",
     "Specific error states for: OAuth failure, timeout, partial data sync, unsupported ERP version. Each error must show: what went wrong, how to fix it, and a path to retry or get help. No generic error pages.",
     "Anirudh", "P0 - Critical", "Onboarding audit complete",
     "2026-02-19", "2026-02-26", "Not Started",
     "Zero generic 'something went wrong' screens", "A failed ERP connection with no recovery path = permanent user loss."],

    ["Phase 0: Pre-Launch", "Product", "Instrument full funnel tracking (10 events)",
     "Track: website visit, signup initiated, signup completed, ERP connection started, ERP connection completed, first Larry query, first chart generated, upgrade prompt shown, upgrade initiated, upgrade completed. Use Mixpanel/Amplitude + Segment.",
     "Anirudh", "P0 - Critical", "None",
     "2026-02-17", "2026-02-24", "Not Started",
     "All 10 events firing correctly in analytics dashboard", "You cannot optimize what you do not measure. Non-negotiable before launch."],

    ["Phase 0: Pre-Launch", "Product", "Instrument engagement metrics",
     "Track: questions per session, questions per day, return rate (Day 1, 3, 7), feature usage (Analyze vs Explain vs Search vs Identify), Close Tracker views.",
     "Anirudh", "P1 - High", "Funnel tracking complete",
     "2026-02-24", "2026-02-28", "Not Started",
     "Engagement dashboard live with daily refresh", "Engagement data drives every decision in May."],

    ["Phase 0: Pre-Launch", "Product", "Build sales handoff trigger notifications",
     "5 triggers auto-notify sales (Rishi initially): (1) User hits 10-question daily limit, (2) User asks about multi-entity, (3) Close Tracker shows bottlenecks, (4) Company matches ICP (Series B+, 10+ team), (5) User asks about Book Close. Slack/email notification with user context.",
     "Anirudh", "P1 - High", "Engagement metrics built",
     "2026-02-28", "2026-03-07", "Not Started",
     "All 5 triggers tested and firing correctly", "Leads without a sales process are wasted. This bridges product to revenue."],

    ["Phase 0: Pre-Launch", "Product", "Build in-product upgrade prompts + flow",
     "When users hit free tier limits (10 questions/day, no downloads, 7-day history), show contextual upgrade modal: what they unlock, price, 1-click upgrade. No redirect to external pricing page. Upgrade must happen in context.",
     "Anirudh", "P1 - High", "Funnel tracking complete",
     "2026-02-26", "2026-03-07", "Not Started",
     "Upgrade flow <3 clicks from limit hit to payment confirmed", "Ambiguity kills self-serve conversion. Make it frictionless."],

    ["Phase 0: Pre-Launch", "Product", "Implement geo-based pricing display",
     "Auto-detect user geography. Show $199/mo for US, $99/mo for India. Never show both prices on same screen. Annual toggle: $199/mo or $1,990/year (2 months free).",
     "Anirudh", "P2 - Medium", "Upgrade flow built",
     "2026-03-07", "2026-03-12", "Not Started",
     "Correct pricing shown based on geo for 10 test locations", "Showing both prices creates confusion and arbitrage."],

    ["Phase 0: Pre-Launch", "Product", "Load test all 4 ERP integrations",
     "Simulate 100+ concurrent ERP connections across Zoho, QBO, ERPNext, Xero. Test: connection time, data sync time, data completeness, error rates. Fix anything that takes >30 min.",
     "Anirudh", "P0 - Critical", "None",
     "2026-03-17", "2026-03-24", "Not Started",
     "All 4 ERPs sync <30 min with <5% error rate at 100 concurrent users", "If this fails at scale on April 1, nothing else matters."],

    ["Phase 0: Pre-Launch", "Product", "Quality-test Larry against 50 real datasets",
     "Run Larry against 50 different ERP data exports. Check: accuracy of insights, appropriateness of suggestions, handling of edge cases (empty data, partial data, unusual COA structures). Document failures.",
     "Anirudh", "P0 - Critical", "Pre-populated queries built",
     "2026-03-10", "2026-03-24", "Not Started",
     "Larry produces accurate answers for >90% of test queries", "Larry giving wrong/misleading answers = trust destroyed permanently."],

    ["Phase 0: Pre-Launch", "Product", "Build Close Tracker basic progress view (Starter tier)",
     "Minimal viable Close Tracker: show % of transactions reconciled, # journal entries posted, visual progress bar. No timeline predictions (Scaling+ only). Position as 'Larry observes your close.'",
     "Anirudh", "P1 - High", "None",
     "2026-02-24", "2026-03-14", "Not Started",
     "Close Tracker shows accurate progress for test accounts", "This is the wedge into paid. It must work at launch."],

    ["Phase 0: Pre-Launch", "Product", "Final QA: 10 test users complete full flow",
     "Recruit 10 people (mix of internal + friendly external). Each must: sign up, connect ERP, ask Larry questions, view Close Tracker, attempt download (hit paywall), complete upgrade. Document every friction point.",
     "Anirudh", "P0 - Critical", "All product tasks above",
     "2026-03-24", "2026-03-28", "Not Started",
     "10/10 users complete flow with zero critical bugs", "Ship-blocking. If this fails, delay launch."],

    # --- Rishi GTM tasks: Week 1-2 ---
    ["PHASE 0", "", "WEEK 1-2: Feb 17-28 — GTM Foundation", "", "", "", "", "", "", "", "", ""],

    ["Phase 0: Pre-Launch", "Marketing", "Rewrite top 20 SEO article titles + meta descriptions",
     "Sort existing blog articles by impressions (Google Search Console). Take top 20. Rewrite titles using benefit+timeframe formula: 'Cut Month-End Close from 10 Days to 3: [Topic] Guide [2026].' Update meta descriptions with result + proof + emotion formula.",
     "Rishi", "P1 - High", "None",
     "2026-02-17", "2026-02-21", "Not Started",
     "20 articles updated; CTR improvement tracked weekly", "Your articles rank but don't get clicks. This is the highest-ROI marketing activity right now."],

    ["Phase 0: Pre-Launch", "Marketing", "LinkedIn posts: Weeks 3-4 of content plan",
     "Continue 3x/week cadence. Posts 7-12 per the content plan: Maker-Checker Framework, One-Person Finance Team Org Chart, Data Entry Waste Calculator, Stale Data Warning System, Institutional Memory, 30-Min Integration Checklist.",
     "Rishi", "P1 - High", "None",
     "2026-02-17", "2026-02-28", "Not Started",
     "6 posts published, engagement rate tracked", "Protect this time ruthlessly. LinkedIn organic is your #1 channel."],

    ["Phase 0: Pre-Launch", "Sales", "Draft sales handoff email templates (5 triggers)",
     "Write personalized outreach templates for each of the 5 sales triggers. Each must reference the user's actual Larry usage. E.g., for question-limit trigger: 'I noticed you've been asking Larry about [X]. Companies like yours typically save [Y] hours with our full platform.'",
     "Rishi", "P1 - High", "None",
     "2026-02-17", "2026-02-21", "Not Started",
     "5 templates drafted and ready for trigger integration", "When a trigger fires, there must be zero delay in knowing what to say."],

    ["Phase 0: Pre-Launch", "Marketing", "Reach out to 10 finance influencers/newsletter operators",
     "Identify 10 people: finance LinkedIn creators, accounting newsletter authors, CFO community leaders. Send personal DM offering early Larry access on their own books. Ask for honest reaction, not a testimonial. Target: 5+ accept.",
     "Rishi", "P1 - High", "None",
     "2026-02-17", "2026-02-24", "Not Started",
     "5+ influencers confirmed for early access", "Authentic reactions from real finance professionals > any ad spend."],

    ["Phase 0: Pre-Launch", "Marketing", "Set up Google AdWords campaigns (do NOT activate yet)",
     "Create campaigns: (1) Branded: 'Komplai,' 'Larry AI finance.' (2) Long-tail: 'MIS automation,' 'AI finance assistant,' 'why does month end close take so long.' (3) Competitor: 'Numeric alternatives,' 'FloQast alternative for startups.' Create ad copy + landing pages for each.",
     "Rishi", "P1 - High", "None",
     "2026-02-24", "2026-03-07", "Not Started",
     "3 campaign groups ready with ads + landing pages", "Do NOT bid on expensive category terms ($20-40/click). Only long-tail and branded."],

    ["Phase 0: Pre-Launch", "Marketing", "Build comparison landing pages (3 pages)",
     "Create: Komplai vs Numeric, Komplai vs FloQast, Komplai vs BlackLine. Use competitive intel from your docs. Honest comparison: where you win, where they win. CTA: 'Try Larry free and see for yourself.'",
     "Rishi", "P2 - Medium", "Google AdWords setup",
     "2026-03-03", "2026-03-10", "Not Started",
     "3 landing pages live with tracked conversion events", "These catch bottom-funnel traffic from competitor search queries."],

    ["Phase 0: Pre-Launch", "Marketing", "Install LinkedIn retargeting pixel on website",
     "Install LinkedIn Insight Tag. Define audience: CFO, Controller, Finance Director, VP Finance job titles. Build retargeting audience from website visitors. Do NOT activate ads yet.",
     "Rishi", "P2 - Medium", "None",
     "2026-02-17", "2026-02-19", "Not Started",
     "Pixel installed, audience accumulating", "Start building the audience NOW. You need 300+ in the pool before retargeting works."],

    ["Phase 0: Pre-Launch", "Hiring", "Hire/contract marketing ops person",
     "Find a part-time marketing ops contractor ($2-3K/month): manage ad campaigns, schedule social posts, set up email sequences, handle initial support tickets. Source from Upwork, LinkedIn, or your network. Must start by March 10.",
     "Rishi", "P0 - Critical", "None",
     "2026-02-17", "2026-03-07", "Not Started",
     "Contractor onboarded and managing ad + email ops", "You CANNOT be CEO, PM, marketer, salesperson, and support agent simultaneously during launch. This is the #1 operational risk."],

    # --- WEEK 3-4: Mar 1-14 ---
    ["PHASE 0", "", "WEEKS 3-4: Mar 1-14 — Build & Test", "", "", "", "", "", "", "", "", ""],

    ["Phase 0: Pre-Launch", "Marketing", "LinkedIn posts: Weeks 5-6 of content plan",
     "Posts 13-18: Instant Answer Workflow, Speed Benchmarks, Execution Plan Template, Retiring Month-End, Automated Accruals Guide, Scaling Without Headcount.",
     "Rishi", "P1 - High", "None",
     "2026-03-02", "2026-03-14", "Not Started",
     "6 posts published, engagement rate tracked", "Keep the cadence. Do not skip."],

    ["Phase 0: Pre-Launch", "Marketing", "Prepare Product Hunt listing",
     "Write tagline, description, 5 screenshots, maker story. Line up a hunter (ideally a known PH community member). Plan first-hour voting: notify your network 24h in advance. Target launch date: April 8-10 (Week 2, not April 1).",
     "Rishi", "P2 - Medium", "None",
     "2026-03-03", "2026-03-14", "Not Started",
     "Full PH listing draft ready, hunter confirmed", "PH in Week 2, not Week 1. Let Week 1 users validate the product first."],

    ["Phase 0: Pre-Launch", "Marketing", "Draft 3-email launch sequence",
     "Email 1 (Apr 1): Launch announcement — what Larry is, what it does, direct link to signup. Email 2 (Apr 3): Feature deep-dive — show Larry in action with real examples. Email 3 (Apr 7): Social proof — early user reactions + CTA.",
     "Rishi", "P1 - High", "None",
     "2026-03-03", "2026-03-10", "Not Started",
     "3 emails drafted, tested, loaded into email platform", "Simple, direct, CTA-focused. No fluff."],

    ["Phase 0: Pre-Launch", "Sales", "Begin partner outreach to accounting firms",
     "Identify 5-10 accounting firms serving Series A-C companies. Pitch: 'We give your clients free Larry access. When they need close automation, you get implementation fees.' Schedule intro calls for April.",
     "Rishi", "P2 - Medium", "None",
     "2026-03-03", "2026-03-14", "Not Started",
     "5 firms contacted, 2-3 intro calls scheduled", "This is the partner-channel revenue stream ($150K-$250K target)."],

    ["Phase 0: Pre-Launch", "Marketing", "Prepare launch day social media plan",
     "Launch day (Apr 1): LinkedIn post (Post 27), Twitter/X post, email blast. Day 2-7: Daily posts with early user reactions, Larry screenshots, engagement prompts. Pre-write all content.",
     "Rishi", "P1 - High", "None",
     "2026-03-10", "2026-03-21", "Not Started",
     "7 days of launch content pre-written and scheduled", "Do not wing launch week. Pre-write everything."],

    # --- WEEKS 5-6: Mar 15-31 ---
    ["PHASE 0", "", "WEEKS 5-6: Mar 15-31 — Final Prep & Soft Launch", "", "", "", "", "", "", "", "", ""],

    ["Phase 0: Pre-Launch", "Marketing", "Execute 'Roast My Ledger' campaign (Post 19)",
     "LinkedIn post: 'Send me your sanitized ledger. I'll run Larry's audit and send you 3 insights you missed.' Commit to 10-20 audits. Each = warm lead + potential case study. Must deliver insights within 24 hours.",
     "Rishi", "P1 - High", "Larry QA against 50 datasets complete",
     "2026-03-17", "2026-03-28", "Not Started",
     "10+ ledger audits completed and shared back", "Your best lead-gen idea. Make it real. 24-hour turnaround is critical."],

    ["Phase 0: Pre-Launch", "Marketing", "LinkedIn Phase 3 content: Weeks 7-8",
     "Posts 19-24: Roast My Ledger, Feature Comparison Matrix, Why We Built It Manifesto, Early Access Fast Track, First Query Prompt, Pricing ROI Calculator. Larry is NOW revealed by name.",
     "Rishi", "P1 - High", "None",
     "2026-03-17", "2026-03-28", "Not Started",
     "6 posts published; waitlist/early access signups tracked", "The reveal phase. Larry goes from mechanism to named product."],

    ["Phase 0: Pre-Launch", "Marketing", "Send early access to confirmed influencers",
     "Give 5+ influencers access to Larry on their own books. Ask them to try it live and share their honest reaction. Provide a simple brief: 'Connect your ERP, ask Larry these 3 questions, tell us what you think.'",
     "Rishi", "P1 - High", "Influencer outreach completed + product QA passed",
     "2026-03-21", "2026-03-28", "Not Started",
     "3+ influencers post or commit to posting about Larry", "Authentic > scripted. Let them find value organically."],

    ["Phase 0: Pre-Launch", "Operations", "Brief launch week support team",
     "Assign roles for April 1-7: (1) Product/bug monitor — Anirudh, (2) Support ticket response <1 hour — marketing ops contractor, (3) Community + social monitoring — Rishi, (4) Backup escalation path. Create shared Slack channel.",
     "Rishi", "P1 - High", "Marketing ops hire complete",
     "2026-03-28", "2026-03-31", "Not Started",
     "All roles assigned with documented escalation path", "If someone tweets 'my ERP connection failed' and no one responds for 6 hours, that's reputational damage."],

    ["Phase 0: Pre-Launch", "Marketing", "LinkedIn launch week content: Week 9",
     "Posts 25-27: Final Checks (Mar 30), Manifesto Redux (Mar 31), MEET LARRY Launch Post (Apr 1). Post 27 must be: problem → solution → CTA (link to signup). No philosophy. Direct path from post to product.",
     "Rishi", "P0 - Critical", "None",
     "2026-03-28", "2026-04-01", "Not Started",
     "Launch post gets 50+ comments and 100+ reactions", "Do NOT bury the CTA. Clear path from post to product."],

    ["Phase 0: Pre-Launch", "Product", "Prepare support/FAQ documentation",
     "Create help docs for: ERP connection (per ERP), common Larry queries, Close Tracker basics, upgrade flow, billing FAQ. Host on help center or knowledge base.",
     "Anirudh", "P1 - High", "Product QA complete",
     "2026-03-21", "2026-03-28", "Not Started",
     "Help docs cover top 20 anticipated questions", "Users who can self-serve support don't become tickets."],

    # ============================================================
    # PHASE 1: LAUNCH MONTH (Apr 1-30)
    # ============================================================
    ["PHASE 1", "", "WEEK 7: Apr 1-7 — LAUNCH WEEK (Controlled)", "", "", "", "", "", "", "", "", ""],

    ["Phase 1: Launch Month", "Marketing", "April 1: Launch post + email blast + social",
     "Execute launch: LinkedIn Post 27 live by 9 AM. Email blast to full list. Twitter/X post. Personal DMs to 20 key people in your network. All pre-written content goes live per schedule.",
     "Rishi", "P0 - Critical", "All Phase 0 complete",
     "2026-04-01", "2026-04-01", "Not Started",
     "Day 1 signups: 50+", "This is the moment. Execute the plan. Do not improvise."],

    ["Phase 1: Launch Month", "Marketing", "Activate Google AdWords (branded terms only)",
     "Turn on branded campaigns only: 'Komplai,' 'Larry AI finance.' Budget: Rs. 10,000/week ($120). Monitor CPC, CTR, conversion to signup. Do NOT activate competitor or category terms yet.",
     "Rishi", "P1 - High", "Launch day",
     "2026-04-01", "2026-04-07", "Not Started",
     "Branded ads live with <$3 CPC and >2% CTR", "Start with branded. Scale to long-tail in Week 2 if branded performs."],

    ["Phase 1: Launch Month", "Product", "Daily monitoring dashboard: Week 1",
     "Every morning at 9 AM, review: signups (total + daily), ERP connections (total + rate), first Larry queries (total + rate), return visits (Day 1 retention). Share dashboard with Rishi. Flag if ERP connection rate <40% or first-query rate <60%.",
     "Anirudh", "P0 - Critical", "Funnel tracking live",
     "2026-04-01", "2026-04-07", "Not Started",
     "Dashboard reviewed daily; issues flagged within 4 hours", "If connection rate <40%, there is a UX problem. Fix before scaling traffic."],

    ["Phase 1: Launch Month", "Sales", "Personal outreach to every Week 1 user",
     "Message every single user who signs up in Week 1. Ask: What brought you here? What did you try first? What worked? What didn't? Use LinkedIn DM, email, or in-product chat. Log responses in spreadsheet.",
     "Rishi", "P0 - Critical", "Launch day",
     "2026-04-01", "2026-04-07", "Not Started",
     "30+ user conversations logged with qualitative feedback", "This qualitative data is worth more than any analytics dashboard."],

    ["Phase 1: Launch Month", "Product", "Fix onboarding friction identified in Week 1",
     "From daily monitoring + user feedback, identify and fix the top 3 friction points. Target <24-hour fix cycle for critical issues. Prioritize: ERP connection failures > Larry accuracy issues > UX confusion.",
     "Anirudh", "P0 - Critical", "Daily monitoring active",
     "2026-04-02", "2026-04-07", "Not Started",
     "Top 3 friction points identified and fixed", "Speed of response in Week 1 determines whether the launch has legs."],

    # --- Week 8: Apr 8-14 ---
    ["PHASE 1", "", "WEEK 8: Apr 8-14 — Iterate & Scale", "", "", "", "", "", "", "", "", ""],

    ["Phase 1: Launch Month", "Marketing", "Product Hunt launch",
     "Launch on PH (target Tuesday or Wednesday). Execute voting plan: notify network 24h before. Rishi: respond to every PH comment within 1 hour. Target: 500+ upvotes, Top 5 of the day.",
     "Rishi", "P1 - High", "Week 1 product issues fixed",
     "2026-04-08", "2026-04-10", "Not Started",
     "500+ upvotes, 200+ signups from PH", "PH can deliver 500-2000 signups in a single day if prepared properly."],

    ["Phase 1: Launch Month", "Marketing", "Scale Google AdWords: add long-tail keywords",
     "Add long-tail campaigns: 'MIS automation,' 'AI finance assistant for CFOs,' 'why does month end close take so long.' Add competitor terms: 'Numeric alternatives.' Budget: Rs. 15,000/week ($180). Create dedicated landing pages per keyword group.",
     "Rishi", "P1 - High", "Branded ads performing",
     "2026-04-08", "2026-04-14", "Not Started",
     "Long-tail campaigns live with <$10 CPC", "Only scale what converts to activated users, not just signups."],

    ["Phase 1: Launch Month", "Marketing", "Shift LinkedIn content: 'What Larry found' posts",
     "Move from educational content to demonstration content. Share anonymized Larry insights from real user data (with permission). Show actual charts, analysis, and answers Larry produced. 3x/week continues.",
     "Rishi", "P1 - High", "Launch complete",
     "2026-04-08", "2026-04-30", "Not Started",
     "3 posts/week showing real Larry outputs", "Show, don't tell. Demonstrated value > theoretical value."],

    ["Phase 1: Launch Month", "Marketing", "Fix existing SEO articles (batch 2: articles 21-50)",
     "After batch 1 (top 20) title rewrites, continue with articles 21-50. Same formula: benefit + timeframe titles, proof + emotion meta descriptions, add schema markup.",
     "Rishi", "P2 - Medium", "Batch 1 done (Feb)",
     "2026-04-08", "2026-04-21", "Not Started",
     "50 total articles optimized; CTR improvement tracked", "SEO compounds. These fixes pay off in May-June."],

    ["Phase 1: Launch Month", "Sales", "First sales outreach from Larry trigger signals",
     "Users who triggered sales handoff notifications get personalized outreach. Reference their actual Larry usage. Goal: start 5+ enterprise conversations in Week 2-3.",
     "Rishi", "P1 - High", "Sales triggers operational",
     "2026-04-08", "2026-04-21", "Not Started",
     "5+ enterprise conversations initiated", "These warm leads are 10x more valuable than cold outbound."],

    # --- Weeks 9-10: Apr 15-30 ---
    ["PHASE 1", "", "WEEKS 9-10: Apr 15-30 — First Conversions", "", "", "", "", "", "", "", "", ""],

    ["Phase 1: Launch Month", "Marketing", "Trigger upgrade email sequence",
     "For users who signed up Weeks 1-2 and used Larry 5+ times: Day 1 email: 'You've asked Larry [X] questions. Scaling gives you unlimited.' Day 3: 'Your team could save [Y] hours.' Day 7: Direct upgrade offer with annual discount.",
     "Rishi", "P0 - Critical", "Upgrade flow built + enough users",
     "2026-04-15", "2026-04-22", "Not Started",
     "Email sequence triggers for all eligible users; 5%+ open-to-upgrade rate", "This is where self-serve revenue starts."],

    ["Phase 1: Launch Month", "Product", "Surface month-end Close Tracker insights proactively",
     "As April month-end approaches (Apr 25-30), push Close Tracker notifications: 'Your April close is X% complete. At current pace, estimated close: Day Y.' Free tier shows basic progress. Scaling shows predictions + bottlenecks.",
     "Anirudh", "P1 - High", "Close Tracker built",
     "2026-04-22", "2026-04-30", "Not Started",
     "Close Tracker notifications sent to all connected users", "Close Tracker is the wedge into paid. Month-end is when it shines."],

    ["Phase 1: Launch Month", "Marketing", "Compile first case study from early adopter",
     "Identify the most engaged user from Week 1-2 signups. Ask for a 15-min call. Document: their problem, how Larry helped, time saved, specific insights. Draft a 1-page case study with before/after metrics.",
     "Rishi", "P1 - High", "Users with 2+ weeks of usage",
     "2026-04-21", "2026-04-30", "Not Started",
     "1 case study draft with real metrics", "You need social proof for May scaling."],

    ["Phase 1: Launch Month", "Marketing", "Activate LinkedIn retargeting ads",
     "Website visitor retargeting with CFO/Controller/Finance Director titles. Budget: Rs. 15,000/month ($180). Creative: 'Your peers are using Larry to close faster. Try it free.' Link to signup.",
     "Rishi", "P2 - Medium", "LinkedIn pixel installed + 300+ in audience pool",
     "2026-04-15", "2026-04-30", "Not Started",
     "Retargeting ads live with tracked conversions", "Only activate if audience pool is 300+. Otherwise, wait."],

    ["Phase 1: Launch Month", "Operations", "Month-end review: April metrics + learnings",
     "Full review: signups, ERP connections, activation rate, questions asked, retention (Day 1/3/7), paid conversions, MRR, enterprise pipeline. Document: what worked, what didn't, what to change for May. Share with team.",
     "Rishi", "P0 - Critical", "Month end",
     "2026-04-28", "2026-04-30", "Not Started",
     "Written review with data tables + decisions for May", "This review drives every May decision. Do not skip."],

    # ============================================================
    # PHASE 2: GROWTH MONTH (May)
    # ============================================================
    ["PHASE 2", "", "PHASE 2: May 1-31 — Growth & Sales Activation", "", "", "", "", "", "", "", "", ""],

    ["Phase 2: Growth Month", "Product", "Launch Growth tier ($499/mo)",
     "Activate Growth tier with: JV Automation (AP/AR), AI-assisted bank recon (Fast Track), weekly validation workflows, email sync. Self-serve with optional onboarding call.",
     "Anirudh", "P0 - Critical", "Phase 1 learnings",
     "2026-05-01", "2026-05-07", "Not Started",
     "Growth tier live and purchasable", "This bridges the 10x gap between Scaling ($199) and Enterprise ($10K+)."],

    ["Phase 2: Growth Month", "Product", "Full Close Tracker features (Scaling+)",
     "Enable timeline predictions and bottleneck analysis for Scaling/Growth users. 'Based on last 6 closes, you finish on Day 8. You're on pace for Day 10.' 'Bank recon is your slowest step.'",
     "Anirudh", "P1 - High", "Basic Close Tracker live",
     "2026-05-01", "2026-05-15", "Not Started",
     "Timeline predictions accurate within 2 days for test accounts", "Full Close Tracker is the primary upgrade driver."],

    ["Phase 2: Growth Month", "Marketing", "Double down on converting channels",
     "Review April data. Identify which channels produced activated users (not just signups). Reallocate May budget: increase spend on channels with >50% activation rate, cut channels with <30% activation rate.",
     "Rishi", "P0 - Critical", "April month-end review",
     "2026-05-01", "2026-05-05", "Not Started",
     "May budget allocated based on April activation data", "A channel that produces signups but low activation is a vanity metric. Kill it."],

    ["Phase 2: Growth Month", "Marketing", "Google AdWords: increase to Rs. 60,000 for May",
     "Add competitor keywords performing from April tests. Create dedicated landing pages for top 3 performing keyword groups. Budget: Rs. 60,000/month ($720).",
     "Rishi", "P1 - High", "April ad performance data",
     "2026-05-05", "2026-05-31", "Not Started",
     "CPA (cost per activated user) tracked and <$50", "Only scale what converts to activated users."],

    ["Phase 2: Growth Month", "Sales", "Warm outbound from Larry usage data",
     "Users who hit question limits, explore multi-entity, or show consistent Close Tracker bottlenecks → personalized outreach referencing their actual usage. Target: 5-8 enterprise conversations in May.",
     "Rishi", "P0 - Critical", "Sales triggers operational",
     "2026-05-05", "2026-05-31", "Not Started",
     "5-8 enterprise conversations; $200K-$400K pipeline", "This is where the real revenue lives. Prioritize over self-serve optimizations."],

    ["Phase 2: Growth Month", "Sales", "Direct outbound to ICP (10 target accounts)",
     "Identify 10 Series A-C companies with 10+ finance teams. Use 'Roast My Ledger' as hook: 'We'll run a free AI audit of your close process.' Personalized emails to CFO/Controller.",
     "Rishi", "P1 - High", "None",
     "2026-05-05", "2026-05-31", "Not Started",
     "10 outbound emails sent; 3+ responses", "Supplement Larry-sourced leads with targeted outbound."],

    ["Phase 2: Growth Month", "Partnerships", "Execute accounting firm partner meetings",
     "Meet with 3-5 firms from March outreach. Present partnership model: free Larry for their clients, implementation fees for close automation. Sign 1-2 partners by May 31.",
     "Rishi", "P1 - High", "March partner outreach",
     "2026-05-05", "2026-05-31", "Not Started",
     "1-2 accounting firm partners signed", "Partner channel = $150K-$250K target. Needs to start producing in June."],

    ["Phase 2: Growth Month", "Marketing", "Community engagement: Reddit, Slack, forums",
     "Weekly participation in r/accounting, r/CFO, finance Slack groups. Answer questions. Share frameworks. Mention Komplai only when relevant. 3-5 substantive contributions per week.",
     "Rishi", "P2 - Medium", "None",
     "2026-05-01", "2026-05-31", "Not Started",
     "15+ community contributions; tracked referral traffic", "Slow burn, high quality. Best leads come from genuine participation."],

    ["Phase 2: Growth Month", "Marketing", "Launch referral program for existing users",
     "Offer existing users 3 months of Scaling free for every referral that converts to paid. Simple referral link/code. Track: referrals sent, signups from referrals, conversions.",
     "Rishi", "P2 - Medium", "Paying users exist",
     "2026-05-15", "2026-05-31", "Not Started",
     "Referral program live; 10+ referrals generated", "Costs nothing but deferred revenue. Your best users are your best salespeople."],

    ["Phase 2: Growth Month", "Marketing", "Publish first case study",
     "Finalize April case study. Publish on website, share on LinkedIn, include in sales outreach. Format: problem → Larry solution → results (with numbers).",
     "Rishi", "P1 - High", "Case study draft from April",
     "2026-05-05", "2026-05-12", "Not Started",
     "Case study published; used in 5+ sales conversations", "Social proof accelerates everything."],

    ["Phase 2: Growth Month", "Operations", "Month-end review: May metrics + learnings",
     "Full review: cumulative signups, paid conversions (Scaling + Growth), MRR, enterprise pipeline, partner progress. Compare to targets. Adjust June plan.",
     "Rishi", "P0 - Critical", "Month end",
     "2026-05-28", "2026-05-31", "Not Started",
     "Written review with revised June targets", "By now you should know if the model works."],

    # ============================================================
    # PHASE 3: ACCELERATION (Jun-Sep)
    # ============================================================
    ["PHASE 3", "", "PHASE 3: Jun-Sep — Acceleration & Enterprise Revenue", "", "", "", "", "", "", "", "", ""],

    ["Phase 3: Acceleration", "Sales", "Close first 2-3 Enterprise deals",
     "Convert April-May pipeline. Target: 2-3 deals at $20K-$30K each. Use case studies, live demos on their data, Larry usage history as proof points.",
     "Rishi", "P0 - Critical", "May pipeline built",
     "2026-06-01", "2026-06-30", "Not Started",
     "$40K-$90K from enterprise in June", "Enterprise is where the revenue math works."],

    ["Phase 3: Acceleration", "Product", "Enterprise tier positioning: full Book Close",
     "Position full Book Close module for Enterprise pipeline. Multi-entity consolidation, maker-checker workflows, custom integrations. Prepare demo flow for Enterprise sales calls.",
     "Anirudh", "P1 - High", "Growth tier stable",
     "2026-06-01", "2026-06-30", "Not Started",
     "Enterprise demo flow documented and tested", "Enterprise deals need a polished story."],

    ["Phase 3: Acceleration", "Marketing", "Scale content + SEO to drive organic traffic",
     "Publish 4-6 new high-intent articles per month using optimized formula. Topics from keyword research: 'close automation for startups,' 'AI reconciliation software,' 'MIS report automation.' SEO should be producing 2,000+ organic visits/month by August.",
     "Rishi", "P1 - High", "SEO fixes from Feb-May",
     "2026-06-01", "2026-09-30", "Not Started",
     "Organic traffic growing 20%+ month-over-month", "Content compounds. June articles pay off in August-September."],

    ["Phase 3: Acceleration", "Sales", "Scaling → Growth upsells",
     "Identify Scaling users who would benefit from Growth features (JV automation, bank recon). Target 10-15% upgrade rate. In-product prompts during close periods: 'Automate this bottleneck with Growth plan.'",
     "Rishi", "P1 - High", "Growth tier live + users on Scaling",
     "2026-06-15", "2026-09-30", "Not Started",
     "10-15% of Scaling users upgrade to Growth", "Expansion revenue is the cheapest revenue."],

    ["Phase 3: Acceleration", "Partnerships", "Partner channel producing referrals",
     "Signed accounting firm partners should be referring clients. Target: 2-3 referral deals per month at $10K-$20K. Provide partners with Larry demo access and sales collateral.",
     "Rishi", "P1 - High", "Partners signed in May",
     "2026-06-15", "2026-09-30", "Not Started",
     "2-3 partner referral deals/month by August", "Partner revenue = $150K-$250K over 6 months. Needs consistent nurturing."],

    ["Phase 3: Acceleration", "Operations", "Monthly revenue reviews (Jun, Jul, Aug, Sep)",
     "Monthly review of: self-serve MRR, enterprise deals closed, partner revenue, total cumulative. Compare to 3-stream model targets. Adjust channels every 2 weeks.",
     "Rishi", "P0 - Critical", "Monthly cadence",
     "2026-06-30", "2026-09-30", "Not Started",
     "Monthly reviews with documented decisions", "Kill what doesn't work. Double down on what does."],
]

# ============================================================
# WRITE TASKS TO SHEET
# ============================================================
row = 5
task_id = 0
for t in tasks:
    phase, category, task, details, owner, priority, dep, start, end, status, metric, notes = t

    # Section header rows
    if phase in ("PHASE 0", "PHASE 1", "PHASE 2", "PHASE 3"):
        ws.merge_cells(f'A{row}:N{row}')
        cell = ws[f'A{row}']
        cell.value = task
        cell.font = Font(name='Arial', bold=True, size=11, color=WHITE)
        if "PHASE 0" in phase:
            cell.fill = PatternFill('solid', fgColor="5C6BC0")
        elif "PHASE 1" in phase:
            cell.fill = PatternFill('solid', fgColor="1976D2")
        elif "PHASE 2" in phase:
            cell.fill = PatternFill('solid', fgColor="388E3C")
        else:
            cell.fill = PatternFill('solid', fgColor=ORANGE)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 24
        row += 1
        continue

    task_id += 1
    ws.row_dimensions[row].height = 60

    # Alternating row color
    if task_id % 2 == 0:
        row_fill = PatternFill('solid', fgColor=LIGHT_GRAY)
    else:
        row_fill = PatternFill('solid', fgColor=WHITE)

    data = [task_id, phase, category, task, details, owner, priority, dep, start, end, "", status, metric, notes]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = body_font
        cell.alignment = body_align if col_idx in (4, 5, 13, 14) else center_align
        cell.border = thin_border
        cell.fill = row_fill

    # Duration formula
    if start and end:
        ws.cell(row=row, column=11).value = f'=J{row}-I{row}+1'
        ws.cell(row=row, column=11).number_format = '0'

    # Date formatting
    for col in (9, 10):
        c = ws.cell(row=row, column=col)
        if c.value:
            c.number_format = 'MMM DD'

    # Priority color coding
    p_cell = ws.cell(row=row, column=7)
    if "P0" in str(priority):
        p_cell.fill = PatternFill('solid', fgColor="FCE4EC")
        p_cell.font = Font(name='Arial', bold=True, size=10, color="C62828")
    elif "P1" in str(priority):
        p_cell.fill = PatternFill('solid', fgColor="FFF8E1")
        p_cell.font = Font(name='Arial', bold=True, size=10, color="F57F17")
    elif "P2" in str(priority):
        p_cell.fill = PatternFill('solid', fgColor="E8F5E9")
        p_cell.font = Font(name='Arial', bold=True, size=10, color="2E7D32")

    # Owner color coding
    o_cell = ws.cell(row=row, column=6)
    if "Anirudh" in str(owner):
        o_cell.fill = PatternFill('solid', fgColor=BLUE_BG)
        o_cell.font = Font(name='Arial', bold=True, size=10, color="1565C0")
    elif "Rishi" in str(owner):
        o_cell.fill = PatternFill('solid', fgColor=ORANGE_BG)
        o_cell.font = Font(name='Arial', bold=True, size=10, color="D84315")

    row += 1

# ============================================================
# SHEET 2: SUMMARY DASHBOARD
# ============================================================
ws2 = wb.create_sheet("Dashboard")
ws2.sheet_properties.tabColor = ORANGE

ws2.merge_cells('A1:F1')
ws2['A1'] = "LAUNCH PLAN DASHBOARD"
ws2['A1'].font = Font(name='Arial', bold=True, size=16, color=NAVY)

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 35

# Task Count Summary
ws2['A3'] = "TASK SUMMARY"
ws2['A3'].font = Font(name='Arial', bold=True, size=12, color=NAVY)

headers2 = ["Category", "Rishi", "Anirudh", "Total", "P0 Tasks", "Key Deadline"]
for i, h in enumerate(headers2, 1):
    c = ws2.cell(row=4, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

summary_data = [
    ["Product/Engineering", 0, 14, 14, 8, "Mar 28: Final QA"],
    ["Marketing", 16, 0, 16, 3, "Apr 1: Launch day"],
    ["Sales", 5, 0, 5, 2, "May 31: Pipeline $200K+"],
    ["Partnerships", 2, 0, 2, 0, "May 31: 1-2 partners signed"],
    ["Operations/Hiring", 4, 0, 4, 3, "Mar 7: Marketing ops hired"],
    ["TOTAL", 27, 14, 41, 16, ""],
]

for r_idx, row_data in enumerate(summary_data, 5):
    for c_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = body_font if r_idx < 10 else Font(name='Arial', bold=True, size=10, color=DARK_GRAY)
        c.alignment = center_align if c_idx < 6 else body_align
        c.border = thin_border
        if r_idx == 10:
            c.fill = PatternFill('solid', fgColor=MED_GRAY)

# Revenue Targets
ws2['A13'] = "REVENUE TARGETS (6-MONTH)"
ws2['A13'].font = Font(name='Arial', bold=True, size=12, color=NAVY)

rev_headers = ["Stream", "Target", "Mechanism", "Confidence", "Key Milestones"]
for i, h in enumerate(rev_headers, 1):
    c = ws2.cell(row=14, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

rev_data = [
    ["Self-serve (Scaling + Growth)", "$150K-$250K", "PLG funnel", "Medium", "200-400 paying subs by Sep"],
    ["Sales-assisted Enterprise", "$400K-$600K", "Larry signals + outbound", "Med-High", "10-15 deals at $20K-$40K"],
    ["Partner / Channel", "$150K-$250K", "Accounting firm referrals", "Low-Med", "2-3 deals/month by Aug"],
    ["TOTAL", "$700K-$1.1M", "", "", "Realistic: $420K-$680K"],
]

for r_idx, row_data in enumerate(rev_data, 15):
    for c_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = body_font if r_idx < 18 else Font(name='Arial', bold=True, size=10, color=DARK_GRAY)
        c.alignment = center_align if c_idx < 5 else body_align
        c.border = thin_border
        if r_idx == 18:
            c.fill = PatternFill('solid', fgColor=MED_GRAY)

# Monthly Milestones
ws2['A21'] = "MONTHLY SUCCESS METRICS"
ws2['A21'].font = Font(name='Arial', bold=True, size=12, color=NAVY)

mile_headers = ["Month", "Signups", "Paid Users", "MRR", "Enterprise Pipeline"]
for i, h in enumerate(mile_headers, 1):
    c = ws2.cell(row=22, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

mile_data = [
    ["April", "200-400", "10-25", "$2K-$7K", "$50K-$100K"],
    ["May", "600-1,000", "40-90", "$8K-$18K", "$200K-$400K"],
    ["June", "1,000-1,600", "80-150", "$15K-$30K", "$300K-$500K"],
    ["July", "1,400-2,200", "120-220", "$22K-$42K", "$400K-$600K"],
    ["August", "1,800-2,800", "160-300", "$30K-$55K", "$500K-$700K"],
    ["September", "2,200-3,500", "200-400", "$40K-$70K", "$600K-$800K"],
]

for r_idx, row_data in enumerate(mile_data, 23):
    for c_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = body_font
        c.alignment = center_align
        c.border = thin_border

# ============================================================
# SHEET 3: BUDGET TRACKER
# ============================================================
ws3 = wb.create_sheet("Budget")
ws3.sheet_properties.tabColor = "388E3C"

ws3.merge_cells('A1:G1')
ws3['A1'] = "MARKETING BUDGET ALLOCATION (Rs. 6,00,000 / ~$7,200)"
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color=NAVY)

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14

budget_headers = ["Channel", "Allocation %", "Total (Rs.)", "Apr", "May", "Jun-Sep", "Notes"]
for i, h in enumerate(budget_headers, 1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

budget_data = [
    ["SEO + Content Optimization", "35%", 210000, 30000, 40000, 140000, "Fix CTR first, then new content"],
    ["Google AdWords (branded + long-tail)", "25%", 150000, 40000, 60000, 50000, "Only long-tail + branded. No expensive category bids."],
    ["Launch Event + Product Hunt", "15%", 90000, 60000, 15000, 15000, "Front-loaded for April launch"],
    ["LinkedIn Retargeting", "15%", 90000, 15000, 25000, 50000, "CFO/Controller titles only. Activate after 300+ pool."],
    ["Community + Partnerships", "10%", 60000, 10000, 20000, 30000, "Event sponsorships, firm outreach, tools"],
    ["TOTAL", "100%", 600000, 155000, 160000, 285000, ""],
]

for r_idx, row_data in enumerate(budget_data, 4):
    for c_idx, val in enumerate(row_data, 1):
        c = ws3.cell(row=r_idx, column=c_idx, value=val)
        c.font = body_font if r_idx < 9 else Font(name='Arial', bold=True, size=10, color=DARK_GRAY)
        c.alignment = center_align if c_idx < 7 else body_align
        c.border = thin_border
        if r_idx == 9:
            c.fill = PatternFill('solid', fgColor=MED_GRAY)
        if c_idx >= 3 and c_idx <= 6 and isinstance(val, (int, float)):
            c.number_format = '#,##0'

# Budget formulas for total row
ws3.cell(row=9, column=3).value = '=SUM(C4:C8)'
ws3.cell(row=9, column=4).value = '=SUM(D4:D8)'
ws3.cell(row=9, column=5).value = '=SUM(E4:E8)'
ws3.cell(row=9, column=6).value = '=SUM(F4:F8)'

for c in range(3, 7):
    ws3.cell(row=9, column=c).number_format = '#,##0'

# ============================================================
# SAVE
# ============================================================
output = "/sessions/funny-stoic-pascal/mnt/komplai-non-technical/Larry-Launch-Action-Plan.xlsx"
wb.save(output)
print(f"Saved: {output}")
